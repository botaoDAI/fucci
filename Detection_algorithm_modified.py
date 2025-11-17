# ================ DESCRIPTION ==============================================================================

# FUCCI 多帧细胞检测器
# 复用原 SEP→LoG→局部极大值参数，对每个 ets 的所有时间帧和前两个荧光通道进行计数
# 支持 RFP 优先的跨通道去重、逐帧进度日志、统计九个位置的均值/方差
# 同时生成可在 FIJI 打开的 *_marked.ome.tif（包含原图与标记遮罩）及 Excel 汇总

# =============== REQUIRED PACKAGES =========================================================================================

import javabridge # To use bioformats 
import bioformats # to open the images
import sep # to determine the background
import Find_Local_Maxima as findMax # python file to determine blobs and local maxima
import pandas # to use dataframes
import sys,os # to access our images and use the terminal
import glob # finds all the pathnames matching a specified pattern according to the rules used by the Unix shell
from tqdm import tqdm # to get a processing bar in the terminal
from scipy.spatial import KDTree
import numpy as np
import openpyxl # for Excel output
import re # for pattern matching
import traceback
import tifffile

# =============== SETTINGS =========================================================================================

# 我们需要指定图像中使用的通道。如果有明场图像和荧光显微镜图像，我们需要将其设置为1！
channel=0
# 为了识别背景，我们定义一个穿过图像的窗口。不应该太小以避免局部效应
bw=256
# 平滑处理
sigma=3.5
# 拉普拉斯阈值
s=0.0375
# 一个spot中被认为是一个spot而不是噪声的最小像素数
ccmin=30
# 如果需要，我们也可以指定定义spot的最大像素数，否则设置为None
ccmax=None
# 如果我们需要在定义的距离内消除重复计数
distance_min=12

# 跨通道去重距离（像素）
merge_distance=3

# 标注输出配置
MARKED_STACK_SUFFIX="_marked"
DEFAULT_MARKED_DIR=os.path.dirname(os.path.abspath(__file__))
ENABLE_MARKED_OUTPUT=False  # 批量运行时可临时改为 False 以跳过生成 OME-TIFF


# =============== HELPER FUNCTIONS =========================================================================================

def _reshape_plane(yy, Nx, Ny, channel_idx):
    """Ensure each channel frame is a 2D contiguous array."""
    if yy.ndim == 3 and yy.shape[2] == 3:
        yy = yy[:, :, channel_idx]
    data = yy.reshape(Ny, Nx)
    return np.ascontiguousarray(data)


def _detect_positions(data, background):
    """Apply the exact SNR -> LoG -> local maxima chain used previously."""
    snr = (data - background.back()) / background.globalrms
    blobs = findMax.getBlobs(snr, s=s, ccmin=ccmin, sigma=sigma, ccmax=ccmax)
    pos = findMax.findMax(blobs, data.shape)
    if pos is None:
        return pandas.DataFrame(columns=["x", "y"])
    return pos


def _df_to_coords(df):
    if df is None or df.empty:
        return np.empty((0, 2), dtype=float)
    return df[["x", "y"]].to_numpy(dtype=float)


def _merge_channel_detections(gfp_df, rfp_df, d_merge):
    gfp_coords = _df_to_coords(gfp_df)
    rfp_coords = _df_to_coords(rfp_df)

    if rfp_coords.size:
        tree = KDTree(rfp_coords)
        mask = np.ones(len(gfp_coords), dtype=bool)
        for idx, point in enumerate(gfp_coords):
            neighbors = tree.query_ball_point(point, d_merge)
            if neighbors:
                mask[idx] = False
        gfp_filtered = gfp_coords[mask]
    else:
        gfp_filtered = gfp_coords

    merged_counts = {
        "channel_0": gfp_filtered.shape[0],
        "channel_1": rfp_coords.shape[0]
    }
    return merged_counts, {0: gfp_filtered, 1: rfp_coords}


def _dtype_limits(dtype):
    if np.issubdtype(dtype, np.integer):
        info = np.iinfo(dtype)
    else:
        info = np.finfo(dtype)
    return info.min, info.max


def _draw_disk(arr, center, radius, value):
    y, x = center
    y = int(round(y))
    x = int(round(x))
    y_min = max(0, y - radius)
    y_max = min(arr.shape[0], y + radius + 1)
    x_min = max(0, x - radius)
    x_max = min(arr.shape[1], x + radius + 1)
    for yi in range(y_min, y_max):
        for xi in range(x_min, x_max):
            if (yi - y) ** 2 + (xi - x) ** 2 <= radius ** 2:
                arr[yi, xi] = value


def _draw_cross(arr, center, arm, value):
    y, x = center
    y = int(round(y))
    x = int(round(x))
    y_min = max(0, y - arm)
    y_max = min(arr.shape[0], y + arm + 1)
    x_min = max(0, x - arm)
    x_max = min(arr.shape[1], x + arm + 1)
    arr[y, x_min:x_max] = value
    arr[y_min:y_max, x] = value


def _create_marker_mask(shape, dtype, coords, marker):
    mask = np.zeros(shape, dtype=dtype)
    if coords.size == 0:
        return mask
    _, max_val = _dtype_limits(dtype)
    if marker == "circle":
        for center in coords:
            cx, cy = center
            _draw_disk(mask, (cy, cx), radius=5, value=max_val)
    else:
        for center in coords:
            cx, cy = center
            _draw_cross(mask, (cy, cx), arm=6, value=max_val)
    return mask


class MarkedStackWriter:
    """Accumulates frames and exports a multi-channel TIFF readable by FIJI."""

    def __init__(self, output_path, frames, total_channels, image_shape, dtype, channel_names):
        self.output_path = output_path
        self.frames = frames
        self.total_channels = total_channels
        self.image_shape = image_shape
        self.dtype = dtype
        self.channel_names = channel_names
        self._buffers = []
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)

    def write(self, frame_planes):
        arr = np.asarray(frame_planes, dtype=self.dtype)
        expected_shape = (self.total_channels, self.image_shape[0], self.image_shape[1])
        if arr.shape != expected_shape:
            raise ValueError(f"帧数据尺寸不匹配: {arr.shape} != {expected_shape}")
        self._buffers.append(arr)

    def close(self):
        if not self._buffers:
            return
        stack = np.stack(self._buffers, axis=0)
        metadata = {
            "axes": "TCYX",
            "frames": self.frames,
            "channels": self.total_channels,
            "channel_names": self.channel_names,
            "hyperstack": True,
            "slices": 1
        }
        tifffile.imwrite(
            self.output_path,
            stack,
            imagej=True,
            metadata=metadata
        )


# =============== MAIN FUNCTION =========================================================================================

def process_folders(base_directories, output_excel_path):
    """
    处理多个文件夹中的frame文件
    
    Args:
        base_directories: 包含要处理文件夹的目录列表
        output_excel_path: Excel输出文件路径
    """
    
    # 初始化结果字典
    results = {}
    
    # 启动Java虚拟机以访问图像
    javabridge.start_vm(class_path=bioformats.JARS)
    
    try:
        for base_dir in base_directories:
            print(f"处理目录: {base_dir}")
            
            # 查找所有包含frame文件的文件夹
            frame_folders = []
            for root, dirs, files in os.walk(base_dir):
                for dir_name in dirs:
                    if dir_name.startswith("_Image_") and dir_name.endswith("__"):
                        frame_path = os.path.join(root, dir_name, "stack1")
                        if os.path.exists(frame_path):
                            # 查找.ets文件
                            ets_files = glob.glob(os.path.join(frame_path, "*.ets"))
                            if ets_files:
                                frame_folders.append((dir_name, ets_files[0]))
            
            print(f"找到 {len(frame_folders)} 个frame文件夹")
            
            # 处理每个frame文件夹
            for folder_name, frame_file in tqdm(frame_folders, desc=f"处理 {os.path.basename(base_dir)}"):
                # 解析puits和位置信息
                # 例如: _Image_A1_01_F98_Fucci__ -> A1, 01
                match = re.search(r'_Image_([A-Z]\d+)_(\d+)_', folder_name)
                if match:
                    puits = match.group(1)  # A1, A2, etc.
                    position = int(match.group(2))  # 01, 02, etc.
                else:
                    print(f"无法解析文件夹名称: {folder_name}")
                    continue
                
                # 处理frame文件
                cell_counts = process_frame_file(frame_file)
                
                # 存储结果
                if puits not in results:
                    results[puits] = {}
                
                results[puits][position] = cell_counts
                
    finally:
        # 关闭Java虚拟机
        javabridge.kill_vm()
    
    # 计算统计信息并输出到Excel
    calculate_and_save_statistics(results, output_excel_path)

def process_frame_file(frame_file, d_merge=merge_distance, marked_output_dir=None):
    """处理单个frame文件，对所有时间帧和前两个通道进行检测并输出标注。"""
    print(f"处理文件: {frame_file}")

    frames_summary = {}
    marked_stack_path = None
    marked_output_dir = marked_output_dir or DEFAULT_MARKED_DIR
    writer = None
    reader = None
    marked_output_path = None

    try:
        ome = bioformats.OMEXML(bioformats.get_omexml_metadata(frame_file))
        nt = ome.image().Pixels.SizeT
        Nx = ome.image().Pixels.SizeX
        Ny = ome.image().Pixels.SizeY
        nchan = ome.image().Pixels.channel_count

        print(f"图像信息: {nt} 帧, 尺寸=({Nx}x{Ny}), {nchan} 个通道")

        if nt == 0:
            print("警告: 没有时间帧")
            return {"channel_0": 0, "channel_1": 0, "frames": frames_summary, "marked_stack": None}

        reader = bioformats.ImageReader(frame_file)
        base_name = os.path.splitext(os.path.basename(frame_file))[0]
        marked_output_path = None
        detection_channels = list(range(min(2, nchan)))
        total_channels = nchan + len(detection_channels)
        channel_names = [f"channel_{idx}_raw" for idx in range(nchan)] + [
            f"channel_{idx}_marks" for idx in detection_channels
        ]
        if ENABLE_MARKED_OUTPUT:
            marked_output_path = os.path.join(marked_output_dir, f"{base_name}{MARKED_STACK_SUFFIX}.ome.tif")

        for frame_idx in range(nt):
            try:
                frame_channel_arrays = {}
                channel_positions = {}
                raw_counts = {}
                bkg = None

                for channel_idx in range(nchan):
                    yy = reader.read(c=channel_idx, t=frame_idx)
                    data = _reshape_plane(yy, Nx, Ny, channel_idx)
                    frame_channel_arrays[channel_idx] = data

                    if channel_idx < min(2, nchan):
                        if channel_idx == 0 or bkg is None:
                            bkg = sep.Background(data, bw=bw, bh=bw)
                        pos = _detect_positions(data, bkg)
                        channel_positions[channel_idx] = pos
                        raw_counts[channel_idx] = len(pos)
                    else:
                        raw_counts[channel_idx] = 0

                merged_counts, merged_coords = _merge_channel_detections(
                    channel_positions.get(0), channel_positions.get(1), d_merge
                )
                merged_total = merged_counts["channel_0"] + merged_counts["channel_1"]
                frames_summary[frame_idx] = {
                    "raw_counts": {
                        "channel_0": raw_counts.get(0, 0),
                        "channel_1": raw_counts.get(1, 0)
                    },
                    "merged_counts": merged_counts,
                    "merged_total": merged_total
                }

                print(
                    f"Frame {frame_idx + 1}/{nt} | raw(GFP,RFP)=({raw_counts.get(0, 0)}, {raw_counts.get(1, 0)}) "
                    f"-> merged(GFP,RFP)=({merged_counts['channel_0']}, {merged_counts['channel_1']}) | 总数={merged_total}"
                )

                if ENABLE_MARKED_OUTPUT:
                    if writer is None:
                        reference_channel = 0 if 0 in frame_channel_arrays else next(iter(frame_channel_arrays))
                        dtype = frame_channel_arrays[reference_channel].dtype
                        writer = MarkedStackWriter(
                            marked_output_path,
                            frames=nt,
                            total_channels=total_channels,
                            image_shape=(Ny, Nx),
                            dtype=dtype,
                            channel_names=channel_names
                        )

                    frame_planes = []
                    for channel_idx in range(nchan):
                        frame_planes.append(frame_channel_arrays[channel_idx])

                    for channel_idx in detection_channels:
                        coords = merged_coords.get(channel_idx, np.empty((0, 2)))
                        marker = "cross" if channel_idx == 0 else "circle"
                        mask = _create_marker_mask(
                            frame_channel_arrays[channel_idx].shape,
                            frame_channel_arrays[channel_idx].dtype,
                            coords,
                            marker
                        )
                        frame_planes.append(mask)

                    writer.write(frame_planes)

            except Exception as frame_err:
                print(f"帧 {frame_idx} 处理异常: {frame_err}")
                traceback.print_exc()
                frames_summary[frame_idx] = {
                    "raw_counts": {"channel_0": 0, "channel_1": 0},
                    "merged_counts": {"channel_0": 0, "channel_1": 0},
                    "merged_total": 0,
                    "error": str(frame_err)
                }
                continue

        if writer is not None:
            marked_stack_path = marked_output_path

        merged_gfp = sum(info["merged_counts"]["channel_0"] for info in frames_summary.values())
        merged_rfp = sum(info["merged_counts"]["channel_1"] for info in frames_summary.values())
        print(
            f"文件汇总 | 帧数: {len(frames_summary)}/{nt} | GFP总计: {merged_gfp} | "
            f"RFP总计: {merged_rfp}"
        )

        last_frame_summary = frames_summary.get(nt - 1, {
            "merged_counts": {"channel_0": 0, "channel_1": 0}
        })
        result = {
            "channel_0": last_frame_summary["merged_counts"].get("channel_0", 0),
            "channel_1": last_frame_summary["merged_counts"].get("channel_1", 0),
            "frames": frames_summary,
            "marked_stack": marked_stack_path
        }
        if ENABLE_MARKED_OUTPUT:
            print(f"标注输出: {marked_stack_path if marked_stack_path else '未生成'}")
        else:
            print("标注输出: 已禁用 (ENABLE_MARKED_OUTPUT=False)")
        return result

    except Exception as e:
        print(f"处理文件 {frame_file} 时出错: {str(e)}")
        traceback.print_exc()
        return {"channel_0": 0, "channel_1": 0, "frames": frames_summary, "marked_stack": None}
    finally:
        if reader is not None:
            reader.close()
        if writer is not None:
            writer.close()

def calculate_and_save_statistics(results, output_excel_path):
    """
    计算每个puits的统计信息并保存到Excel
    
    Args:
        results: 包含所有结果的字典
        output_excel_path: Excel输出文件路径
    """
    print("计算统计信息...")
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    
    # 为每个通道创建工作表
    for channel_idx in range(2):
        ws = wb.create_sheet(title=f"Channel_{channel_idx}")
        
        # 设置标题行
        headers = ["Puits", "Position_01", "Position_02", "Position_03", "Position_04", 
                  "Position_05", "Position_06", "Position_07", "Position_08", "Position_09", 
                  "Mean", "Std"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 填充数据
        row = 2
        for puits in sorted(results.keys()):
            puits_data = results[puits]
            
            # 收集九个位置的数据
            position_counts = []
            for pos in range(1, 10):  # 01到09
                if pos in puits_data:
                    count = puits_data[pos].get(f"channel_{channel_idx}", 0)
                    position_counts.append(count)
                else:
                    position_counts.append(0)
            
            # 计算平均值和标准差
            if position_counts:
                mean_val = np.mean(position_counts)
                std_val = np.std(position_counts, ddof=1)  # 样本标准差
            else:
                mean_val = 0
                std_val = 0
            
            # 写入数据
            ws.cell(row=row, column=1, value=puits)
            for col, count in enumerate(position_counts, 2):
                ws.cell(row=row, column=col, value=count)
            ws.cell(row=row, column=11, value=mean_val)
            ws.cell(row=row, column=12, value=std_val)
            
            row += 1
    
    # 删除默认工作表
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # 保存Excel文件
    wb.save(output_excel_path)
    print(f"结果已保存到: {output_excel_path}")

# =============== MAIN EXECUTION =========================================================================================

if __name__ == "__main__":
    # 设置要处理的目录
    base_directories = [
        "/Users/dai/Desktop/fucci/20251003 f98 fucci-1",
#        "/Users/dai/Desktop/fucci/20251003 f98 fucci-1 last line"
#        "test1"
    ]
    
    # 设置输出Excel文件路径
    output_excel_path = "/Users/dai/Desktop/fucci/cell_count_results-test.xlsx"
    
    print("开始处理细胞计数...")
    print(f"处理目录: {base_directories}")
    print(f"输出文件: {output_excel_path}")
    
    # 确认执行
    ans = input("是否继续执行? (y/n): ")
    if ans.lower() != "y":
        print("程序已取消")
        sys.exit()
    
    # 执行处理
    process_folders(base_directories, output_excel_path)
    print("处理完成！")
